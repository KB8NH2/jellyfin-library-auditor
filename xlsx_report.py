"""Combined Excel workbook generation for audit results.

Writes one worksheet per audited server, each holding the same rows as that
server's own audit CSV in a named Excel Table, plus an optional "diffs"
worksheet when two servers were compared. Yes/No columns on a server sheet
get a yellow background for "Yes" cells; diff cells (any "left|right" value
where the two sides differ) get the same treatment on the diffs sheet. The
Episode column is formatted as text (rather than CSV's leading-apostrophe
guard) so a merged range like "19-20" isn't misread as a date. Every cell,
header and data alike, is wrapped and center-aligned, and every column after
Title has a fixed width rather than the content-fitted width the columns up
through Title get. A server sheet also gets a "Problems" column (a per-row
count of that row's "Yes" cells) and a "Totals" row below the table (a
per-column count of "Yes" cells, plus the sum of "Problems") - neither
exists in the server's own audit CSV, which stays a plain per-item export.

Each server also gets a second, smaller "<label> Series Summary" worksheet:
one row per TV series (movies excluded - there's nothing to roll up), with
"Number of Seasons"/"Total Number of Episodes" holding how many distinct
seasons and episode files that series has locally (context for the counts
next to them, not a problem indicator on their own), "Missing
Seasons"/"Missing Episodes" holding how many individual seasons/episodes
are missing across that series (summed from audit.missing_number_count()
on that series' missing_seasons/missing_episodes finding message(s), since
that's the only place the count is recorded - see that function's
docstring), "Missing Subtitles" holding how many of the series' episodes
lack one, and a "Complete" column that's "Yes" only when all three missing
counts are 0 - collapsing a per-episode Yes/No sprawl to one row per
series, so a fully-clean series (every numbered season and episode
present, every episode with an English subtitle track) is a single glance
instead of a per-episode scan.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from audit import missing_number_count
from audit_types import AuditFinding
import compare_csv_files
from models import MediaItem
from output_layout import audit_results_xlsx_path
from reports.generator import CSV_HEADER
from reports.generator import _csv_rows
from results import AuditServerResult


DIFFS_SHEET_LABEL = "diffs"
PROBLEMS_COLUMN_LABEL = "Problems"
TOTALS_ROW_LABEL = "Totals"
SERIES_SUMMARY_SHEET_SUFFIX = "Series Summary"
COMPLETE_COLUMN_LABEL = "Complete"
SERIES_SUMMARY_HEADER = (
    "Library",
    "Series",
    "Number of Seasons",
    "Total Number of Episodes",
    "Missing Seasons",
    "Missing Episodes",
    "Missing Subtitles",
    COMPLETE_COLUMN_LABEL,
)

# Purely informational counts (how big the series is) - summed in the
# Totals row like the problem counts below, but never highlighted, since a
# high count here isn't a problem to fix.
_SERIES_SUMMARY_INFO_COUNT_COLUMNS = frozenset({"Number of Seasons", "Total Number of Episodes"})
# How much of the series is missing - highlighted yellow when greater than
# 0, same as a server sheet's own Yes/No columns.
_SERIES_SUMMARY_PROBLEM_COUNT_COLUMNS = frozenset(
    {"Missing Seasons", "Missing Episodes", "Missing Subtitles"}
)
_SERVER_IDENTITY_COLUMNS = frozenset(
    {"Library", "Base Directory", "Base Filename", "Series", "Title", "Season", "Episode"}
)
_YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
_INVALID_SHEET_NAME_CHARS = re.compile(r"[:\\/?*\[\]]")
_INVALID_TABLE_NAME_CHARS = re.compile(r"[^A-Za-z0-9_]+")
_MAX_SHEET_NAME_LENGTH = 31
_MAX_COLUMN_WIDTH = 60
_FIXED_COLUMN_WIDTH_AFTER_TITLE = 12
_TEXT_NUMBER_FORMAT = "@"
_WRAP_CENTER_ALIGNMENT = Alignment(wrap_text=True, horizontal="center", vertical="center")
_TOTALS_ROW_FONT = Font(bold=True)
# Column after which every subsequent column (Season, Episode, and all the
# Yes/No/Problems columns) gets a fixed width instead of the generic
# content-fitted one - see _set_fixed_width_after_title(). Both CSV_HEADER
# and the diffs sheet's header (compare_csv_files.build_header(CSV_HEADER))
# keep "Title" at the same position, so one column name lookup covers both
# sheet shapes.
_FIXED_WIDTH_START_COLUMN = "Title"


def write_audit_results_workbook(
    root_dir: Path,
    server_results: tuple[AuditServerResult, ...],
    *,
    diff_results: tuple[AuditServerResult, AuditServerResult] | None = None,
) -> Path:
    """Write the combined audit-results Excel workbook to ``root_dir``.

    Args:
        root_dir: Audit-results root directory (same directory as the
            top-level index.html).
        server_results: One worksheet is written per entry, named after the
            server and containing the same rows as that server's own audit
            CSV (see reports.generator.write_csv_report).
        diff_results: An optional (left, right) pair of already-audited
            server results to diff into a "diffs" worksheet, in the same
            row shape compare_csv_files.py produces for its CSV output.

    Returns:
        The written workbook's path.

    Raises:
        PermissionError: If ``output_path`` couldn't be overwritten - most
            often because it's currently open in Excel or another program,
            which locks the file. Re-raised with a clearer message; the
            previous run's xlsx is left in place on disk in that case, even
            though this run's CSV/HTML reports (written earlier) did get
            updated - the two can briefly disagree until the xlsx write
            succeeds on a later run.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    used_sheet_titles: set[str] = set()
    used_table_names: set[str] = set()

    for result in server_results:
        label = _server_label(result)
        rows = _csv_rows(result)
        _write_server_sheet(
            workbook,
            label=label,
            rows=rows,
            used_sheet_titles=used_sheet_titles,
            used_table_names=used_table_names,
        )
        _write_series_summary_sheet(
            workbook,
            label=label,
            items=result.audited_items,
            findings=result.findings,
            used_sheet_titles=used_sheet_titles,
            used_table_names=used_table_names,
        )

    if diff_results is not None:
        left_result, right_result = diff_results
        _write_diffs_sheet(
            workbook,
            left_result=left_result,
            right_result=right_result,
            used_sheet_titles=used_sheet_titles,
            used_table_names=used_table_names,
        )

    output_path = audit_results_xlsx_path(root_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
    except PermissionError as error:
        raise PermissionError(
            f"Could not write {output_path} - it may be open in Excel or another "
            "program. Close it and re-run. The CSV/HTML reports were written "
            "successfully and reflect this run; only the previous run's xlsx "
            "file is left in place."
        ) from error
    return output_path


def _write_server_sheet(
    workbook: Workbook,
    *,
    label: str,
    rows: tuple[tuple[str, ...], ...],
    used_sheet_titles: set[str],
    used_table_names: set[str],
) -> None:
    """Write one server's audit rows to a new named-table worksheet."""
    rows = compare_csv_files.strip_episode_guard(CSV_HEADER, rows)
    sheet = workbook.create_sheet(_unique_sheet_title(label, used_sheet_titles))

    yes_no_indices = _yes_no_column_indices(CSV_HEADER)
    first_yes_no_letter = get_column_letter(yes_no_indices[0] + 1)
    last_yes_no_letter = get_column_letter(yes_no_indices[-1] + 1)
    header = CSV_HEADER + (PROBLEMS_COLUMN_LABEL,)
    rows_with_problems = tuple(
        row + (f'=COUNTIF({first_yes_no_letter}{row_number}:{last_yes_no_letter}{row_number},"Yes")',)
        for row_number, row in enumerate(rows, start=2)
    )

    last_row = _write_table(
        sheet,
        header=header,
        rows=rows_with_problems,
        table_name=_unique_table_name(label, used_table_names),
    )
    _set_fixed_width_after_title(sheet, header)

    _add_yes_no_conditional_formatting(sheet, CSV_HEADER, last_row)
    _add_episode_text_format(sheet, CSV_HEADER, last_row)
    _add_wrap_center_alignment(sheet, header, last_row)
    _add_totals_row(
        sheet,
        yes_no_indices=yes_no_indices,
        problems_column_index=len(CSV_HEADER),
        last_row=last_row,
    )


def _write_series_summary_sheet(
    workbook: Workbook,
    *,
    label: str,
    items: tuple[MediaItem, ...],
    findings: tuple[AuditFinding, ...],
    used_sheet_titles: set[str],
    used_table_names: set[str],
) -> None:
    """Write one server's per-series completeness rollup to a new worksheet."""
    summary_rows = _series_summary_rows(items, findings)
    sheet_label = f"{label} {SERIES_SUMMARY_SHEET_SUFFIX}"
    sheet = workbook.create_sheet(_unique_sheet_title(sheet_label, used_sheet_titles))

    last_row = _write_table(
        sheet,
        header=SERIES_SUMMARY_HEADER,
        rows=summary_rows,
        table_name=_unique_table_name(sheet_label, used_table_names),
    )
    _set_fixed_width_after_title(sheet, SERIES_SUMMARY_HEADER, start_column="Series")

    _add_series_summary_conditional_formatting(sheet, SERIES_SUMMARY_HEADER, last_row)
    _add_wrap_center_alignment(sheet, SERIES_SUMMARY_HEADER, last_row)
    _add_series_summary_totals_row(sheet, SERIES_SUMMARY_HEADER, last_row)


def _series_summary_rows(
    items: tuple[MediaItem, ...],
    findings: tuple[AuditFinding, ...],
) -> tuple[tuple[object, ...], ...]:
    """Return one row per TV series, rolled up from its episodes and findings.

    "Number of Seasons"/"Total Number of Episodes" hold how many distinct
    seasons and how many episode files that series has locally - context for
    the counts next to them, not problem indicators on their own (a big
    show isn't a "problem" the way a missing season is).

    "Missing Seasons"/"Missing Episodes" hold how many individual
    seasons/episodes are missing for that series - summed from
    audit.missing_number_count() on that series' missing_seasons/
    missing_episodes finding message(s) (missing_episodes can produce more
    than one finding per series, one per season with a gap; missing_seasons
    produces at most one). "Missing Subtitles" holds how many of the
    series' episodes are missing English subtitles - one
    missing_english_subtitles finding per affected episode, simply counted.
    "Complete" reads "Yes" only when all three missing counts are 0 - every
    numbered season and episode is present on disk, and every one of those
    episodes has an English subtitle track.

    Every TV series with at least one local episode gets a row, even with
    all-zero counts - a fully clean series must still show up as
    "Complete", not be silently absent. Movies (no series name) contribute
    nothing - there are no seasons/episodes to be complete or incomplete
    about.

    Args:
        items: Every media item audited for one server (movies and
            episodes alike) - see results.AuditServerResult.audited_items.
        findings: Every finding produced for that same server - see
            results.AuditServerResult.findings.
    """
    season_numbers: dict[tuple[str, str], set[int]] = {}
    episode_counts: dict[tuple[str, str], int] = {}
    for item in items:
        if not item.is_episode or not item.series_name:
            continue
        key = (item.library, item.series_name)
        season_numbers.setdefault(key, set())
        if item.season_number is not None:
            season_numbers[key].add(item.season_number)
        episode_counts[key] = episode_counts.get(key, 0) + 1

    missing_seasons_counts: dict[tuple[str, str], int] = {}
    missing_episodes_counts: dict[tuple[str, str], int] = {}
    missing_subtitles_counts: dict[tuple[str, str], int] = {}
    for finding in findings:
        series_name = finding.media_item.series_name
        if not series_name:
            continue
        key = (finding.media_item.library, series_name)
        if finding.check_name == "missing_seasons":
            missing_seasons_counts[key] = missing_seasons_counts.get(
                key, 0
            ) + (missing_number_count(finding.message) or 0)
        elif finding.check_name == "missing_episodes":
            missing_episodes_counts[key] = missing_episodes_counts.get(
                key, 0
            ) + (missing_number_count(finding.message) or 0)
        elif finding.check_name == "missing_english_subtitles":
            missing_subtitles_counts[key] = missing_subtitles_counts.get(key, 0) + 1

    summary_rows = []
    for library, series_name in sorted(
        season_numbers, key=lambda key: (key[0].casefold(), key[1].casefold())
    ):
        key = (library, series_name)
        missing_seasons = missing_seasons_counts.get(key, 0)
        missing_episodes = missing_episodes_counts.get(key, 0)
        missing_subtitles = missing_subtitles_counts.get(key, 0)
        complete = missing_seasons == 0 and missing_episodes == 0 and missing_subtitles == 0
        summary_rows.append(
            (
                library,
                series_name,
                len(season_numbers[key]),
                episode_counts[key],
                missing_seasons,
                missing_episodes,
                missing_subtitles,
                "Yes" if complete else "No",
            )
        )
    return tuple(summary_rows)


def _add_series_summary_conditional_formatting(
    sheet: Worksheet, header: tuple[str, ...], last_row: int
) -> None:
    """Highlight the series summary sheet's problem cells.

    "Missing Seasons"/"Missing Episodes"/"Missing Subtitles" get a yellow
    background when their count is greater than 0 (a problem to fix);
    "Complete" gets the opposite - a green background on "Yes" - since
    there it's the desirable outcome, not a problem. "Number of
    Seasons"/"Total Number of Episodes" are purely informational and never
    highlighted - a big show isn't a problem.
    """
    if last_row < 2:
        return
    for index, column_name in enumerate(header):
        column_letter = get_column_letter(index + 1)
        if column_name in _SERIES_SUMMARY_PROBLEM_COUNT_COLUMNS:
            sheet.conditional_formatting.add(
                f"{column_letter}2:{column_letter}{last_row}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=_YELLOW_FILL),
            )
        elif column_name == COMPLETE_COLUMN_LABEL:
            sheet.conditional_formatting.add(
                f"{column_letter}2:{column_letter}{last_row}",
                CellIsRule(operator="equal", formula=['"Yes"'], fill=_GREEN_FILL),
            )


def _add_series_summary_totals_row(
    sheet: Worksheet, header: tuple[str, ...], last_row: int
) -> None:
    """Write a "Totals" row directly below the series summary table.

    Every count column - "Number of Seasons"/"Total Number of Episodes" as
    well as "Missing Seasons"/"Missing Episodes"/"Missing Subtitles" - gets
    the sum of its per-series counts; "Complete" gets a count of its own
    "Yes" cells, same as a server sheet's Yes/No columns. Deliberately left
    out of the table's own ref, same reason as _add_totals_row(): a row
    inside the table would participate in the table's own sort/filter,
    which could otherwise scatter this row away from the bottom of the
    sheet.
    """
    totals_row = last_row + 1
    sheet.cell(row=totals_row, column=1, value=TOTALS_ROW_LABEL)

    has_data_rows = last_row >= 2
    count_columns = _SERIES_SUMMARY_INFO_COUNT_COLUMNS | _SERIES_SUMMARY_PROBLEM_COUNT_COLUMNS
    for index, column_name in enumerate(header):
        column_letter = get_column_letter(index + 1)
        if column_name in count_columns:
            value = (
                f"=SUM({column_letter}2:{column_letter}{last_row})" if has_data_rows else 0
            )
            sheet.cell(row=totals_row, column=index + 1, value=value)
        elif column_name == COMPLETE_COLUMN_LABEL:
            value = (
                f'=COUNTIF({column_letter}2:{column_letter}{last_row},"Yes")'
                if has_data_rows
                else 0
            )
            sheet.cell(row=totals_row, column=index + 1, value=value)

    for cell in sheet[totals_row]:
        cell.font = _TOTALS_ROW_FONT
        cell.alignment = _WRAP_CENTER_ALIGNMENT


def _write_diffs_sheet(
    workbook: Workbook,
    *,
    left_result: AuditServerResult,
    right_result: AuditServerResult,
    used_sheet_titles: set[str],
    used_table_names: set[str],
) -> None:
    """Write the left/right diff rows to a new "diffs" worksheet."""
    left_rows = compare_csv_files.strip_episode_guard(CSV_HEADER, _csv_rows(left_result))
    right_rows = compare_csv_files.strip_episode_guard(CSV_HEADER, _csv_rows(right_result))
    diff_header, diff_rows = compare_csv_files.diff_header_and_rows(
        CSV_HEADER, left_rows, CSV_HEADER, right_rows
    )
    # diff_header_and_rows() re-applies its own CSV text guard (a leading
    # apostrophe) to a non-differing Episode value, matching
    # compare_csv_files.py's own CSV output - not wanted here, since the
    # Episode column instead gets a real text number format below.
    diff_rows = compare_csv_files.strip_episode_guard(diff_header, diff_rows)

    sheet = workbook.create_sheet(_unique_sheet_title(DIFFS_SHEET_LABEL, used_sheet_titles))
    last_row = _write_table(
        sheet,
        header=diff_header,
        rows=diff_rows,
        table_name=_unique_table_name(DIFFS_SHEET_LABEL, used_table_names),
    )
    _set_fixed_width_after_title(sheet, diff_header)

    _add_diff_conditional_formatting(sheet, diff_header, last_row)
    _add_episode_text_format(sheet, diff_header, last_row)
    _add_wrap_center_alignment(sheet, diff_header, last_row)


def _write_table(
    sheet: Worksheet,
    *,
    header: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
    table_name: str,
) -> int:
    """Write a header/rows table to ``sheet`` as a named Excel Table.

    Returns the 1-indexed row number of the last written row (the header
    row when ``rows`` is empty).
    """
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"

    last_row = len(rows) + 1
    last_column_letter = get_column_letter(len(header))
    table = Table(displayName=table_name, ref=f"A1:{last_column_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sheet.add_table(table)

    for index, column_name in enumerate(header):
        column_letter = get_column_letter(index + 1)
        sheet.column_dimensions[column_letter].width = _column_width(column_name, rows, index)

    return last_row


def _yes_no_column_indices(header: tuple[str, ...]) -> tuple[int, ...]:
    """Return the 0-indexed positions of every non-identity (Yes/No) column."""
    return tuple(index for index, name in enumerate(header) if name not in _SERVER_IDENTITY_COLUMNS)


def _add_totals_row(
    sheet: Worksheet,
    *,
    yes_no_indices: tuple[int, ...],
    problems_column_index: int,
    last_row: int,
) -> None:
    """Write a "Totals" row directly below the table.

    Each Yes/No column gets a count of its own "Yes" cells; the Problems
    column gets the sum of its per-row counts (so it equals the same total
    either way). Deliberately left out of the table's own ref - a row
    inside the table would participate in the table's own sort/filter,
    which could otherwise scatter this row away from the bottom of the
    sheet.
    """
    totals_row = last_row + 1
    sheet.cell(row=totals_row, column=1, value=TOTALS_ROW_LABEL)

    # With no data rows (last_row == 1, just the header), the totals row is
    # row 2 - the same row a "2:{last_row}" formula would reference, which
    # would make the cell reference itself. Write a literal 0 instead in
    # that case rather than a self-referential formula.
    has_data_rows = last_row >= 2

    for index in yes_no_indices:
        column_letter = get_column_letter(index + 1)
        value = (
            f'=COUNTIF({column_letter}2:{column_letter}{last_row},"Yes")' if has_data_rows else 0
        )
        sheet.cell(row=totals_row, column=index + 1, value=value)

    problems_column_letter = get_column_letter(problems_column_index + 1)
    problems_value = (
        f"=SUM({problems_column_letter}2:{problems_column_letter}{last_row})"
        if has_data_rows
        else 0
    )
    sheet.cell(row=totals_row, column=problems_column_index + 1, value=problems_value)

    for cell in sheet[totals_row]:
        cell.font = _TOTALS_ROW_FONT
        cell.alignment = _WRAP_CENTER_ALIGNMENT


def _add_yes_no_conditional_formatting(
    sheet: Worksheet, header: tuple[str, ...], last_row: int
) -> None:
    """Highlight "Yes" cells in every non-identity column with a yellow background."""
    if last_row < 2:
        return
    for index in _yes_no_column_indices(header):
        column_letter = get_column_letter(index + 1)
        sheet.conditional_formatting.add(
            f"{column_letter}2:{column_letter}{last_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=_YELLOW_FILL),
        )


def _add_diff_conditional_formatting(
    sheet: Worksheet, header: tuple[str, ...], last_row: int
) -> None:
    """Highlight any cell containing "left|right" values that differ.

    A single formula rule covers every data cell: it fires when the cell
    contains "|" and the text before it doesn't match the text after it,
    which is exactly compare_csv_files.py's own definition of a differing
    column (an identity column only ever shows "|" when the two sides
    differ; a test-criteria column always shows "left|right", so this
    still only lights up where "left" and "right" actually disagree).
    """
    if last_row < 2:
        return
    last_column_letter = get_column_letter(len(header))
    formula = (
        'AND(ISNUMBER(SEARCH("|",A2)),'
        'LEFT(A2,FIND("|",A2)-1)<>MID(A2,FIND("|",A2)+1,LEN(A2)))'
    )
    sheet.conditional_formatting.add(
        f"A2:{last_column_letter}{last_row}",
        FormulaRule(formula=[formula], fill=_YELLOW_FILL),
    )


def _add_episode_text_format(sheet: Worksheet, header: tuple[str, ...], last_row: int) -> None:
    """Format the Episode column's data cells as text.

    Excel's own auto-detection would otherwise read a merged-episode value
    like "19-20" as a date; a real text number format prevents that without
    needing compare_csv_files.py's CSV-only leading-apostrophe guard (which
    is stripped from these rows before they're written - see
    compare_csv_files.strip_episode_guard()).
    """
    if "Episode" not in header or last_row < 2:
        return
    column_letter = get_column_letter(header.index("Episode") + 1)
    for row in range(2, last_row + 1):
        sheet[f"{column_letter}{row}"].number_format = _TEXT_NUMBER_FORMAT


def _add_wrap_center_alignment(sheet: Worksheet, header: tuple[str, ...], last_row: int) -> None:
    """Wrap and center every cell in the table, header and data alike."""
    for index in range(len(header)):
        column_letter = get_column_letter(index + 1)
        for row in range(1, last_row + 1):
            sheet[f"{column_letter}{row}"].alignment = _WRAP_CENTER_ALIGNMENT


def _set_fixed_width_after_title(
    sheet: Worksheet, header: tuple[str, ...], *, start_column: str = _FIXED_WIDTH_START_COLUMN
) -> None:
    """Set every column after ``start_column`` to a fixed width.

    Overrides the generic content-fitted width _write_table() gives every
    column by default - a Yes/No or short numeric column doesn't need
    content-fit sizing, and a fixed width keeps those columns visually
    consistent regardless of what happens to be in them. Columns up through
    ``start_column`` (Library, Base Directory, Base Filename, Series, Title
    on a server/diffs sheet; Library, Series on the series summary sheet,
    passed explicitly since it has no Title column of its own) keep their
    content-fitted width, since those can genuinely vary a lot in length.
    """
    if start_column not in header:
        return
    start_index = header.index(start_column) + 1
    for index in range(start_index, len(header)):
        column_letter = get_column_letter(index + 1)
        sheet.column_dimensions[column_letter].width = _FIXED_COLUMN_WIDTH_AFTER_TITLE


def _server_label(result: AuditServerResult) -> str:
    """Return the display label used for one server's sheet/table name."""
    return result.server_name or result.server_key or result.server_url or "Server"


def _unique_sheet_title(label: str, used_titles: set[str]) -> str:
    """Return an Excel-legal, workbook-unique worksheet title derived from ``label``."""
    base = _INVALID_SHEET_NAME_CHARS.sub("_", label.strip())[:_MAX_SHEET_NAME_LENGTH] or "Sheet"
    title = base
    suffix = 2
    while title.casefold() in used_titles:
        suffix_text = f" ({suffix})"
        title = base[: _MAX_SHEET_NAME_LENGTH - len(suffix_text)] + suffix_text
        suffix += 1
    used_titles.add(title.casefold())
    return title


def _unique_table_name(label: str, used_names: set[str]) -> str:
    """Return an Excel-legal, workbook-unique table name derived from ``label``."""
    slug = _INVALID_TABLE_NAME_CHARS.sub("_", label.strip()).strip("_") or "Table"
    if slug[0].isdigit():
        slug = f"T_{slug}"
    name = slug
    suffix = 2
    while name.casefold() in used_names:
        name = f"{slug}_{suffix}"
        suffix += 1
    used_names.add(name.casefold())
    return name


def _column_width(column_name: str, rows: tuple[tuple[object, ...], ...], index: int) -> float:
    """Return a content-fitted column width, capped so long paths stay usable.

    ``str()``-converts each cell before measuring rather than assuming a
    string, since the series summary sheet's count columns hold real ``int``
    values rather than pre-formatted text.
    """
    longest = max((len(str(row[index])) for row in rows), default=0)
    return min(max(len(column_name), longest) + 2, _MAX_COLUMN_WIDTH)


__all__ = ["write_audit_results_workbook"]
