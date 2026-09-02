"""Tests for xlsx_report.py."""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

import openpyxl

import xlsx_report
from audit_types import AuditCategory
from audit_types import AuditSeverity
from reports.generator import CSV_HEADER

from tests.helpers import _make_finding
from tests.helpers import _make_item
from tests.helpers import _make_single_library_result


class WriteAuditResultsWorkbookTests(unittest.TestCase):
    def test_title_column_matches_the_items_current_title(self) -> None:
        """Regression test: the workbook's Title column must reflect each
        item's current title (the same value the CSV/HTML reports show),
        not a stale or independently-derived one - the server sheet is
        built from the exact same rows reports.generator._csv_rows()
        produces for the CSV.
        """
        item = _make_item("Correct New Title", is_movie=True, is_episode=False)
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary"]
            header = [cell.value for cell in sheet[1]]
            title_column = header.index("Title") + 1

            self.assertEqual(sheet.cell(row=2, column=title_column).value, "Correct New Title")

    def test_path_is_split_into_base_directory_and_base_filename_columns(self) -> None:
        item = _make_item(
            "Movie One",
            is_movie=True,
            is_episode=False,
            path=Path("Movies/Movie One (2024)/Movie One (2024).mkv"),
        )
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary"]
            header = [cell.value for cell in sheet[1]]

        self.assertNotIn("Path", header)
        base_directory_column = header.index("Base Directory") + 1
        base_filename_column = header.index("Base Filename") + 1
        self.assertEqual(sheet.cell(row=2, column=base_directory_column).value, "Movie One (2024)")
        self.assertEqual(
            sheet.cell(row=2, column=base_filename_column).value, "Movie One (2024).mkv"
        )

    def test_permission_error_on_save_is_reraised_with_a_clearer_message(self) -> None:
        """Regression test: if the xlsx is open in Excel (or another program
        holding it locked), Workbook.save() raises a plain PermissionError
        with no indication of why - that must be surfaced as an actionable
        message instead of a bare stack trace, since the previous run's
        stale file is left in place on disk when this happens."""
        item = _make_item("Some Title")
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            with patch.object(
                xlsx_report.Workbook, "save", side_effect=PermissionError("locked")
            ):
                with self.assertRaises(PermissionError) as context:
                    xlsx_report.write_audit_results_workbook(root_dir, (result,))

        self.assertIn("open in Excel", str(context.exception))

    def test_every_cell_is_wrapped_and_center_aligned(self) -> None:
        item = _make_item("Some Title", library="Movies")
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary"]
            header = [cell.value for cell in sheet[1]]

        # "Library" is the very first column - before Title, and well
        # before the old "Season onward" cutoff - confirming wrap/center
        # now applies to every column, not just a trailing subset.
        library_column = header.index("Library") + 1
        for row in (1, 2):
            cell = sheet.cell(row=row, column=library_column)
            self.assertTrue(cell.alignment.wrap_text)
            self.assertEqual(cell.alignment.horizontal, "center")
            self.assertEqual(cell.alignment.vertical, "center")

    def test_columns_after_title_have_a_fixed_width(self) -> None:
        item = _make_item(
            "A Title Much Longer Than Eleven Characters",
            library="Movies",
        )
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary"]
            header = [cell.value for cell in sheet[1]]

        title_index = header.index("Title")
        # Title itself keeps its content-fitted width, which for this
        # deliberately long title is wider than the fixed width every
        # later column gets.
        title_letter = openpyxl.utils.get_column_letter(title_index + 1)
        self.assertGreater(
            sheet.column_dimensions[title_letter].width, xlsx_report._FIXED_COLUMN_WIDTH_AFTER_TITLE
        )
        for index in range(title_index + 1, len(header)):
            column_letter = openpyxl.utils.get_column_letter(index + 1)
            self.assertEqual(
                sheet.column_dimensions[column_letter].width,
                xlsx_report._FIXED_COLUMN_WIDTH_AFTER_TITLE,
            )

    def test_diffs_sheet_also_gets_wrap_center_and_fixed_width(self) -> None:
        left_item = _make_item("Same Title", library="Movies")
        right_item = _make_item("Same Title", library="Movies")
        left_result = _make_single_library_result(
            (left_item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Left",
            server_key="left",
        )
        right_result = _make_single_library_result(
            (right_item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Right",
            server_key="right",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(
                root_dir, (left_result, right_result), diff_results=(left_result, right_result)
            )

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["diffs"]
            header = [cell.value for cell in sheet[1]]

        library_column = header.index("Library") + 1
        self.assertTrue(sheet.cell(row=1, column=library_column).alignment.wrap_text)
        title_index = header.index("Title")
        after_title_letter = openpyxl.utils.get_column_letter(title_index + 2)
        self.assertEqual(
            sheet.column_dimensions[after_title_letter].width,
            xlsx_report._FIXED_COLUMN_WIDTH_AFTER_TITLE,
        )

    def test_header_matches_csv_header_plus_problems_column(self) -> None:
        item = _make_item("Some Title")
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary"]
            header = tuple(cell.value for cell in sheet[1])

        self.assertEqual(header, CSV_HEADER + ("Problems",))


class SeriesSummarySheetTests(unittest.TestCase):
    def test_marks_a_series_complete_when_nothing_is_missing(self) -> None:
        item = _make_item(
            "Ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Complete Show",
            season_number=1,
            episode_number=1,
        )
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            self.assertIn("Primary Series Summary", workbook.sheetnames)
            sheet = workbook["Primary Series Summary"]
            header = tuple(cell.value for cell in sheet[1])
            row = tuple(cell.value for cell in sheet[2])

        self.assertEqual(
            header,
            (
                "Library",
                "Series",
                "Number of Seasons",
                "Total Number of Episodes",
                "Missing Seasons",
                "Missing Episodes",
                "Missing Subtitles",
                "Complete",
            ),
        )
        self.assertEqual(row, ("TV Shows", "Complete Show", 1, 1, 0, 0, 0, "Yes"))

    def test_marks_a_series_incomplete_when_any_episode_is_missing_subtitles(self) -> None:
        """Regression test: "Missing Subtitles" counts how many episodes of
        the series are missing English subtitles - here 1 of 2 - and the
        series reads "Complete" = "No" as soon as that count is nonzero,
        even when no seasons/episodes are missing.
        """
        subtitled_episode = _make_item(
            "Ep1",
            item_id="ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Partial Show",
            season_number=1,
            episode_number=1,
        )
        unsubtitled_episode = _make_item(
            "Ep2",
            item_id="ep2",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Partial Show",
            season_number=1,
            episode_number=2,
        )
        finding = _make_finding(
            category=AuditCategory.SUBTITLES,
            severity=AuditSeverity.WARNING,
            title="Ep2",
            check_name="missing_english_subtitles",
            media_item=unsubtitled_episode,
        )
        result = _make_single_library_result(
            (subtitled_episode, unsubtitled_episode),
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
            findings=(finding,),
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]
            rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2, max_row=2)]

        self.assertEqual(rows, [("TV Shows", "Partial Show", 1, 2, 0, 0, 1, "No")])

    def test_marks_a_series_incomplete_when_a_season_is_missing(self) -> None:
        """Regression test: "Missing Seasons" holds the actual count of
        missing seasons (2, 4, and 6 - three individual seasons, even though
        the range collapses to two comma-separated segments in the
        message), not just a Yes/No flag.
        """
        item = _make_item(
            "Ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Show With A Gap",
            season_number=1,
            episode_number=1,
        )
        finding = _make_finding(
            category=AuditCategory.METADATA,
            severity=AuditSeverity.WARNING,
            title="Ep1",
            check_name="missing_seasons",
            message="Missing seasons: 2, 4-6, out of 7 seasons.",
            media_item=item,
        )
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
            findings=(finding,),
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]
            row = tuple(cell.value for cell in sheet[2])

        self.assertEqual(row, ("TV Shows", "Show With A Gap", 1, 1, 4, 0, 0, "No"))

    def test_sums_missing_episodes_across_multiple_seasons_of_one_series(self) -> None:
        """Regression test: missing_tv_season_episodes() produces one
        finding per season with a gap, so a series missing episodes in two
        different seasons has two missing_episodes findings - "Missing
        Episodes" must sum both, not just reflect the last one seen.
        """
        item = _make_item(
            "Ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Show With Two Gaps",
            season_number=1,
            episode_number=1,
        )
        season_one_finding = _make_finding(
            category=AuditCategory.METADATA,
            severity=AuditSeverity.WARNING,
            title="Ep1",
            check_name="missing_episodes",
            message="Missing episodes: 2, out of 3 episodes.",
            media_item=item,
        )
        season_two_finding = _make_finding(
            category=AuditCategory.METADATA,
            severity=AuditSeverity.WARNING,
            title="Ep1",
            check_name="missing_episodes",
            message="Missing episodes: 3-4, out of 4 episodes.",
            media_item=item,
        )
        result = _make_single_library_result(
            (item,),
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
            findings=(season_one_finding, season_two_finding),
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]
            row = tuple(cell.value for cell in sheet[2])

        self.assertEqual(row, ("TV Shows", "Show With Two Gaps", 1, 1, 0, 3, 0, "No"))

    def test_number_of_seasons_and_episodes_reflect_distinct_local_counts(self) -> None:
        """Regression test: "Number of Seasons" counts distinct season
        numbers present locally (2, here - seasons 1 and 2 - not 5, the
        number of episode files), and "Total Number of Episodes" counts
        every local episode file for the series, regardless of any missing
        seasons/episodes/subtitles findings."""
        items = tuple(
            _make_item(
                f"S1E{number}",
                item_id=f"s1e{number}",
                is_movie=False,
                is_episode=True,
                library="TV Shows",
                series_name="Big Show",
                season_number=1,
                episode_number=number,
            )
            for number in range(1, 4)
        ) + tuple(
            _make_item(
                f"S2E{number}",
                item_id=f"s2e{number}",
                is_movie=False,
                is_episode=True,
                library="TV Shows",
                series_name="Big Show",
                season_number=2,
                episode_number=number,
            )
            for number in range(1, 3)
        )
        result = _make_single_library_result(
            items,
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]
            row = tuple(cell.value for cell in sheet[2])

        self.assertEqual(row, ("TV Shows", "Big Show", 2, 5, 0, 0, 0, "Yes"))

    def test_excludes_movies_from_the_summary(self) -> None:
        movie = _make_item("A Movie", is_movie=True, is_episode=False, library="Movies")
        result = _make_single_library_result(
            (movie,),
            library_id="lib1",
            library_name="Movies",
            collection_type="movies",
            server_name="Primary",
            server_key="primary",
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]

        self.assertEqual(sheet.max_row, 2)  # header row + Totals row only

    def test_totals_row_counts_incomplete_and_complete_series(self) -> None:
        complete_item = _make_item(
            "Ep1",
            item_id="complete-ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Complete Show",
            season_number=1,
            episode_number=1,
        )
        incomplete_item = _make_item(
            "Ep1",
            item_id="incomplete-ep1",
            is_movie=False,
            is_episode=True,
            library="TV Shows",
            series_name="Incomplete Show",
            season_number=1,
            episode_number=1,
        )
        finding = _make_finding(
            category=AuditCategory.SUBTITLES,
            severity=AuditSeverity.WARNING,
            title="Ep1",
            check_name="missing_english_subtitles",
            media_item=incomplete_item,
        )
        result = _make_single_library_result(
            (complete_item, incomplete_item),
            library_id="lib1",
            library_name="TV Shows",
            collection_type="tvshows",
            server_name="Primary",
            server_key="primary",
            findings=(finding,),
        )

        with TemporaryDirectory() as temp_dir:
            root_dir = Path(temp_dir) / "audit_results"
            output_path = xlsx_report.write_audit_results_workbook(root_dir, (result,))

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["Primary Series Summary"]
            totals_row = tuple(cell.value for cell in sheet[4])

        self.assertEqual(totals_row[0], "Totals")
        header = xlsx_report.SERIES_SUMMARY_HEADER
        complete_letter = openpyxl.utils.get_column_letter(header.index("Complete") + 1)
        subtitles_letter = openpyxl.utils.get_column_letter(header.index("Missing Subtitles") + 1)
        self.assertEqual(
            totals_row[header.index("Complete")], f'=COUNTIF({complete_letter}2:{complete_letter}3,"Yes")'
        )
        self.assertEqual(
            totals_row[header.index("Missing Subtitles")],
            f"=SUM({subtitles_letter}2:{subtitles_letter}3)",
        )


if __name__ == "__main__":
    unittest.main()
